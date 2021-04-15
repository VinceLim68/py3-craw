# coding:utf-8
import traceback
import datetime
import time
import re
import pymysql,openpyxl,xlrd
from openpyxl import Workbook,load_workbook

# 统一拿连接数据库的实例，以后有修改就可以一次改完
def get_database() -> object:
    return pymysql.connect(host ='localhost', user ="root", passwd ="root", db ="property_info", charset ="utf8", port = 3306)

def mylog(func):
    def _deco(*a, **b):
        try:
            # result = func(*a,**b) 
            # func(*a,**b)                  #以前只有return,但发现有返回值的函数加装饰器时，总是返回空，所以改成return func(*a,**b)
            return func(*a, **b)
            # return
        except Exception as e:
            # result = None
            with open('logtest.txt', 'a+') as fout:
                fout.write('\n               *******' + func.__name__ + '*******,error record by @mylog on ' + str(
                    datetime.datetime.now()) + '*************\n')
                traceback.print_exc(file=fout)
                print(traceback.format_exc())
                # return func
                # return result

    return _deco


def exeTime(func):
    def newFunc(*args, **args2):
        t0 = time.time()
        back = func(*args, **args2)
        print("@ %.3fs taken for {%s}\n" % (time.time() - t0, func.__name__))
        return back

    return newFunc


def pri(string):
    if len(string) % 2 != 0:
        string = string + ' '
    print(string)


def clearStr(string):
    # 清理字符串中的回车、空格等
    if isinstance(string, str):
        string = string.replace('<br/>', '').replace('\r', '').replace(' ', '').replace('\n', '').strip()
        string = string.replace(u'\xa0', '')  # 去除&nbsp;
        string = string.replace('<b>', '').replace('</b>', '')
        string = string.replace('<', '').replace('>', '')
        # 去除图形符号
        try:
            co = re.compile(u'[\U00010000-\U0010ffff]')
        except re.error:
            co = re.compile(u'[\uD800-\uDBFF][\uDC00-\uDFFF]')
        string = co.sub(u'', string)
        # # string = re.sub("[•.]", "", string)
        # string = re.sub('[&*()+-`~!@#$%{}\':;,^\\\\.<>《》/?！￥…（）•—【】‘；：”“’。，、？]', "", string)
        string = re.sub("[&`~()!@#$%{}\';,<>《》?！￥…（）•—【】‘；：”“’。，、？]", "", string)
    return string


def ShowInvalideData(each_data):
    # temp = {}
    if each_data:
        if not each_data.has_key('total_floor'):  # 2016.6.1搜房网老是出现无效数据，进行判断，发现是别墅没有记载楼层信息造成的
            if u"别墅" in each_data['title']:
                each_data['total_floor'] = 4
                each_data['floor_index'] = 1
                each_data['spatial_arrangement'] = each_data['spatial_arrangement'] + u"别墅"
                # page_datas.append(each_data)
                print("没有总楼层,按照别墅填充！！！")
                for key, value in each_data.items(): print(" %s : %s" % (key, value))
                print("=" * 35)
                return True
            else:
                print("！！！没有总楼层！！！")
        elif not each_data.has_key('total_price'):
            print("！！！没有总价！！！")
        elif not each_data.has_key('area'):
            print("！！！没有面积！！！")
        elif not each_data.has_key('community_name'):
            print("！！！小区名！！！")
        printDic(each_data)
        print("=" * 35)
    return False


def confir(str):
    for i in range(0, 32):
        str = str.replace(chr(i), '')
    return str


def printDic(data):
    if isinstance(data, dict):
        for key in data:
            print('%20s : %s' % (key, data[key]))
    else:
        # data = to_str(data)
        print(data)


def strToInt(string1):
    if isinstance(string1, str):
        b = re.findall(r'\d+\.?\d*', string1)
        # b = re.findall(r'\d+.\d+', string1)
        # print(b)
        try:
            # string1 = int(round(float(string1)))
            string1 = int(round(float(b[0])))
            # string1 =
            # string1 = int(round(float(re.sub("\D", "",string1))))
        # except ValueError as e:
        except:
            string1 = 0
    return string1

# 输出列表
def priList(list_name, level=0):
    i = 1
    for yuansu in list_name:
        if isinstance(yuansu, list):  # 判断当前元素是不是列表
            priList(yuansu, level + 1)  # 如是,则递归调用,并且标记当前元素是列表
        else:
            for tab in range(level):  # 固定次数
                print("\t", end='')
            printDic(yuansu)
            print('************************  {0}  *************************'.format(i))
            i += 1


def to_str(bytes_or_str):
    if isinstance(bytes_or_str, bytes):
        value = bytes_or_str.decode('utf-8')
    else:
        value = bytes_or_str
    return value  # Instance of str

def clear_comm(str):
    #清除小区名称中（）内的部分
    return str.split('(')[0].split('（')[0].strip()

# 把中文字符串转成阿拉伯数字
CN_NUM = {
    '〇' : 0, '一' : 1, '二' : 2, '三' : 3, '四' : 4, '五' : 5, '六' : 6, '七' : 7, '八' : 8, '九' : 9, '零' : 0,
    '壹' : 1, '贰' : 2, '叁' : 3, '肆' : 4, '伍' : 5, '陆' : 6, '柒' : 7, '捌' : 8, '玖' : 9, '貮' : 2, '两' : 2,
}

CN_UNIT = {
    '十' : 10,
    '拾' : 10,
    '百' : 100,
    '佰' : 100,
    '千' : 1000,
    '仟' : 1000,
    '万' : 10000,
    '萬' : 10000,
    '亿' : 100000000,
    '億' : 100000000,
    '兆' : 1000000000000,
}

def chinese_to_arabic(cn:str) -> int:
    unit = 0   # current
    ldig = []  # digest
    for cndig in reversed(cn):
        if cndig in CN_UNIT:
            unit = CN_UNIT.get(cndig)
            if unit == 10000 or unit == 100000000:
                ldig.append(unit)
                unit = 1
        else:
            dig = CN_NUM.get(cndig)
            if unit:
                dig *= unit
                unit = 0
            ldig.append(dig)
    if unit == 10:
        ldig.append(10)
    val, tmp = 0, 0
    for x in reversed(ldig):
        if x == 10000 or x == 100000000:
            val += tmp * x
            tmp = 0
        else:
            tmp += x
    val += tmp
    return val

def cn2num(cn:str):
    # import re
    pattern = re.compile(r'[一二三四五六七八九十百千万亿零]+')
    print(pattern.findall(cn))

# TODO: make a full unittest
def test():
    test_dig = ['二二百',
                '一十一',
                '一百二十三',
                '一千二百零三',
                '一万一千一百零一',
                '十万零三千六百零九',
                '一百二十三万四千五百六十七',
                '一千一百二十三万四千五百六十七',
                '一亿一千一百二十三万四千五百六十七',
                '一百零二亿五千零一万零一千零三十八']
    for cn in test_dig:
        x = chinese_to_arabic(cn)
        print(cn, x)
    assert x == 10250011038

# 编译拆分地址的正则
def Break_up_the_address(address):
    address = str(address)
    regex = dict()
    regex['province'] = "(?P<province>[^省]+?自治区|.+?省)"
    regex['city'] = "(?P<city>^[^市]{,5}?市|.+?地区|.+?盟|.+?自治州)"
    regex['county'] = "(?P<county>.+?县|.+?旗|.+?乡|^[^镇小道路街巷]+?区)"
    # regex['county'] = "(?P<county>.+?县|.+?旗|.+?乡|.+?[^小]区)"
    regex['town'] = "(?P<town>[^县]+?镇|.+?街道办事处|.+?街道|.+?基地|.+?农场|.+?林场|.+?种场|.+?牧场|.+?殖场|.+?苏木)"
    # regex['road'] = "(?P<road>.+?路|.+?道|.+?街|.+?巷|.+?road|.+?线|.+?段|.+?里|.+?弄|.+?条|.+?出口|.+?入口|.+?高速|.+?快速|.+?胡同)"
    # regex['road_number'] = "(?P<road_number>([a-zA-Z0-9一二三四五六七八九十百千甲乙丙丁支-]+(弄|号院|号)(?!楼))+)"
    # regex['road'] = "(?P<road>.+?路|.+?道|.+?街|.+?巷|.+?road|.+?线|.+?段|.+?里|.+?弄|.+?条|.+?出口|.+?入口|.+?高速|.+?快速|.+?胡同)" \
    #                 "(?P<road_number>([a-zA-Z0-9一二三四五六七八九十百千甲之乙丙丁支-]+(弄|号院|号)(?!楼)))"
    regex['road'] = "(?P<road>^\D+?)" \
                    "(?P<road_number>([a-zA-Z0-9一二三四五六七八九十百千甲之乙丙丁支-]+(号楼|号院|号|弄)))"
    regex['period'] = "(?P<period>[a-zA-Z0-9一二三四五六七八九十百千东西南北中-]+[区期])"
    regex['building'] = "(?P<building>第?[-a-zA-Z0-9一二三四五六七八九十百千]+([#幢楼栋座梯棟门]+|号楼))"
    regex['danyuan'] = "(?P<danyuan>附?[a-zA-Z0-9一二三四五六七八九十百千号#-]{1,3}(梯|座|单元))"
    regex['ceng'] = "(?P<ceng>[-0-9一二三四五六七八九十百千负地第下上]+层)"
    regex['room'] = "(?P<room>[-0-9一二三四五六七八九十百千第负]*[室房号单元]*)(?P<other>.*)"
    address_dic = {}
    for key, value in regex.items():
        match = re.search(value,address)
        if match:
            match_dic = match.groupdict()
            for k1,v1 in match_dic.items():
                address = address.replace(match_dic[k1],"",1)
            address_dic.update(match_dic)
        else:
            address_dic[key] = None
    address_dic['community_name'] = address
    return address_dic

# 把{}保存为xlsx
def dic2Excel(to_file,in_dic):
    wb = Workbook()
    ws1 = wb.create_sheet('result')
    row = 1
    for k,v in in_dic.items():
        ws1.cell(row, 1).value = k
        ws1.cell(row, 2).value = v
        row += 1
    wb.save(to_file)
    print("结果表已经生成，路径为", to_file)

#[(),()]保存为excel
def list2Excel(to_file,input_list,title="结果"):
    wb = Workbook()
    ws = wb.create_sheet(title,0)
    for item in input_list:
        ws.append(item)
    wb.save(to_file)
    print("结果表已经生成，路径为", to_file)

# [{},{},{}]模式保存为xlsx
def saveExcel(mypath, mydata,sheet="Sheet1"):
    try:
        wb = load_workbook(mypath)
    except Exception as e:
        wb = Workbook()

    ws1 = wb.create_sheet(sheet, 0)
    # ws1 = wb.active
    # ws1.title = "result"
    col_names = list(mydata[0].keys())
    row = 1
    for header in col_names:
        col_index = col_names.index(header) + 1
        ws1.cell(row,col_index).value = header
    row += 1
    for data_row in mydata:
        for _key, _value in data_row.items():
            if _key not in col_names:
                col_names.append(_key)
                ws1.cell(1, col_names.index(_key) + 1).value = _key
            col = col_names.index(_key) + 1
            ws1.cell(row, col).value = _value
        row += 1
    wb.save(mypath)
    print("结果表已经生成，路径为", mypath)

# excel=>[{},{},{}]
def read_excel(mypath, sheetname="Sheet1"):
    #打开excel表，填写路径
    book = xlrd.open_workbook(mypath)
    #找到sheet页
    table = book.sheet_by_name(sheetname)
    #获取总行数总列数
    row_Num = table.nrows
    col_Num = table.ncols

    s =[]
    key =table.row_values(0)# 这是第一行数据，作为字典的key值

    if row_Num <= 1:
        print("没数据")
    else:
        j = 1
        for i in range(row_Num-1):
            d ={}
            values = table.row_values(j)
            for x in range(col_Num):
                # 把key值对应的value赋值给key，每行循环
                d[key[x]]=values[x]
            j+=1
            # 把字典加到列表中
            # print(d)
            s.append(d)
        return s

# 输入文件，指定地址字段名，批量拆解地址，返回一个excel
def batch_breakup_address(infile,address_col_name,outfile,sheet="Sheet1"):
    get_addresses = read_excel(infile,sheet)
    result_data = []
    for address in get_addresses:
        address_string = clearStr(address[address_col_name])
        # address_string = (address[address_col_name])
        # parentheses_string = get_text_in_parentheses(address_string)    #取出括号里的值
        # address_string = address_string.replace(parentheses_string, '')     #把地址中带括号的去除
        small_address = Break_up_the_address(address_string)
        # print(small_address)
        d = {**address, **small_address}
        result_data.append(d)
    saveExcel(outfile,result_data)

# 取出括号里的值
def get_text_in_parentheses(text):
    # 先把括号替换一下
    # text = re.sub("[（【\[{]", "(", text)
    # text = re.sub("[）】\]}]", ")", text)
    # print(text)
    regex = r"(?P<inParentheses>[（【\[{\(].+[）】\]}\)])"
    pattern = re.compile(regex)
    match = re.search(pattern,text)
    get_str = ""
    if match:
        get_str = match.group(0)
    return get_str

# print(get_text_in_parentheses("鼓楼区鼓东街道鼓东路创业公寓{创业新村]1#楼2层10附属间、603单元"))
# print(clearStr("晋安区新店镇福飞北路189号福晟.钱隆御景（二期）7#楼2205单元"))
# batch_breakup_address("C:\\Users\\15007\\Desktop\\input.xls","押品详细地址","C:\\Users\\15007\\Desktop\\拆分地址.xlsx","SQL Results")

# print(s)
# saveExcel("test.xlsx",s)
# export_excel(s)
# a1 = (type(s[0]))
# print(a1)
# for key in a1.keys:
#     print(key)
# print(Break_up_the_address('上海市普陀区宜川六村51号A座207-8室'))
# print(Break_up_the_address('上海市建设南路2345弄234号345号楼1-8室'))
# print(Break_up_the_address('递铺镇迎宾大道187号天平花园北区（凯迪大厦）22幢801室'))
# print(Break_up_the_address('重庆市南岸区南湖支路19号1幢1层6号门市'))
# printDic(Break_up_the_address('（二拍）宜兴市张渚镇聚贤山庄23幢503'))
# printDic(Break_up_the_address('宜兴市解放东路1号的房地产'))
# printDic(Break_up_the_address('后江埭路46-4号之7'))
# printDic(Break_up_the_address('新胜村554-556号'))
# printDic(Break_up_the_address('渝西大道支路25号'))
# printDic(Break_up_the_address('天文大道奥园城市天地B区'))
# printDic(Break_up_the_address('无锡市湖滨壹号花园3-3203号的房产——232.55㎡'))
# printDic(Break_up_the_address('长兴雉城大西门清文路52号校区内3幢401室（含车库及装修）'))
# printDic(Break_up_the_address('厦门市思明区湖滨南路146-150号地下一层第09号车位'))
# print(Break_up_the_address('北京昌平区陈家营西路3号院10号楼3层2单元302'))
# print(Break_up_the_address('玉屏山长富花园华富楼12层1201室'))
# printDic(Break_up_the_address('西陂镇龙腾中路618号（五洲财富）4幢13层1324（楼中楼）'))
# printDic(Break_up_the_address('重庆市大足区棠香街道办事处五星大道368号2幢3-3号住宅'))
# printDic(Break_up_the_address('重庆市万州区土堡街300号附2号1单元8-2室房屋拍卖'))

# str2 = 'morning公馆308公安局集资房MORNING公馆V8区'
# a = strToInt(str2)
# print(a)
# print(clearStr(str2))
# if isinstance(str2,str):
#     print(clearStr(str2))
# print(str2)
# print(clearStr(str2))
# test()
# (cn2num('二二百安商业街一千零七十九号小区'))


