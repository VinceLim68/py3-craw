import re
import time

import openpyxl
import openpyxl as op


def read_cmnt():
    # 创建一个工作簿
    # wb = openpyxl.Workbook()
    # 创建一个test_case的sheet表单
    # wb.create_sheet('test_case')
    # 保存为一个xlsx格式的文件
    # wb.save('cases.xlsx')
    # 读取excel中的数据
    # 第一步：打开工作簿
    wb = openpyxl.load_workbook(r'template_excel/广州市数据_20210203.xlsx')
    # wb = openpyxl.load_workbook(r'template_excel/重庆上库基价.xlsx')
    # wb = openpyxl.load_workbook(r'source_file/福州20210121新增基准价表(1)(1).xlsx')
    # 第二步：选取表单
    sh = wb['sheet1']
    # 第三步：读取数据
    # 参数 row:行  column：列
    ce = sh.cell(row=1, column=2)  # 读取第一行，第一列的数据
    # print(ce.value)
    # 按行读取数据 list(sh.rows)
    # print(list(sh.rows)[1:])  # 按行读取数据，去掉第一行的表头信息数据
    area_list = []
    cmnt_list = []
    adress_list = []
    all_list = []
    for cases in list(sh.rows)[1:]:
        # case_id = cases[0].value
        case_area = cases[0].value
        case_cmnt = cases[1].value
        case_adress = cases[2].value
        # case_excepted2 = cases[4].value
        # case_data = cases[2].value
        # print(case_excepted)
        area_list.append(case_area)
        cmnt_list.append(case_cmnt)
        adress_list.append(case_adress)
        # adress_list.append(case_excepted2)
    # print('area_list----------', area_list)
    # print('cmnt_list----------', cmnt_list)
    # print('adress_list---------', adress_list)
    all_list.append(area_list)
    all_list.append(cmnt_list)
    all_list.append(adress_list)
    # print(all_list)
    # 关闭工作薄
    wb.close()
    return all_list


def read_list():
    all_list = read_cmnt()
    # print(all_list[0])
    loudong = []
    for area, cmnt, address in zip(all_list[0], all_list[1], all_list[2]):
        cmnt = str(cmnt).replace('(', '').replace(')', '').replace('（', '').replace('）', '').rstrip(';')
        cmnt = list(set(re.split('[;|]', cmnt)))
        cmnt = ';'.join(cmnt)
        # print(area)
        # print(cmnt)
        # print(address)
        # loudong.append(area)
        try:
            # with open('loudong.txt', 'a+')as f:
            #     f.write(area + ' ' + str(cmnt) + ' ' + address + '\n')
            if str(address).find(';') != -1 or str(address).find('|') != -1 or str(address).find('、') != -1 or str(address).find('.') != -1:
                address = str(address).rstrip(';')
                address_list = list(re.split('[;|.、]', address))
                ad_pre = ''
                for ad in address_list:
                    reg_fuhao = re.search('^\D+', ad)
                    if reg_fuhao:
                        ad_pre = reg_fuhao.group()
                        ad = ad
                        address_reg1 = re.search('(\d+)号?楼?[-～－~至]+(\d+)号?楼?', ad)
                        address_reg2 = re.search('(\d+)号?楼?[-～－~至]+(\d+)号?楼?[-～－~至]+(\d+)号?楼?', ad)
                        if address_reg2:
                            with open('loudong.txt', 'a+')as f:
                                f.write(area + ' ' + cmnt + ' ' + ad + '\n')

                        elif address_reg1:
                            n1 = int(address_reg1.group(1))
                            n2 = int(address_reg1.group(2))
                            if n1 < n2:
                                if (n1 % 2 == 0 and n2 % 2 == 0) or (n1 % 2 != 0 and n2 % 2 != 0):
                                    for i in range(n1, n2 + 1, 2):
                                        ad = re.sub('(\d+)号?楼?[-～－~至]+(\d+)号?楼?', '', ad)
                                        ad1 = ad + str(i) + '号'
                                        with open('loudong.txt', 'a+')as f:
                                            f.write(area + ' ' + cmnt + ' ' + ad1 + '\n')
                                else:
                                    for i in range(n1, n2 + 1):
                                        ad = re.sub('(\d+)号?楼?[-～－~至]+(\d+)号?楼?', '', ad)
                                        ad1 = ad + str(i) + '号'
                                        with open('loudong.txt', 'a+')as f:
                                            f.write(area + ' ' + cmnt + ' ' + ad1 + '\n')
                            else:
                                with open('loudong.txt', 'a+')as f:
                                    f.write(area + ' ' + cmnt + ' ' + ad + '\n')


                        else:
                            with open('loudong.txt', 'a+')as f:
                                f.write(area + ' ' + cmnt + ' ' + ad + '\n')


                    else:
                        ad = ad_pre + ad
                        if ad.endswith('号'):
                            address_reg1 = re.search('(\d+)号?楼?[-～－~至]+(\d+)号?楼?', ad)
                            address_reg2 = re.search('(\d+)号?楼?[-～－~至]+(\d+)号?楼?[-～－~至]+(\d+)号?楼?', ad)
                            if address_reg2:
                                with open('loudong.txt', 'a+')as f:
                                    f.write(area + ' ' + cmnt + ' ' + ad + '\n')

                            elif address_reg1:
                                n1 = int(address_reg1.group(1))
                                n2 = int(address_reg1.group(2))
                                if n1 < n2:
                                    if (n1 % 2 == 0 and n2 % 2 == 0) or (n1 % 2 != 0 and n2 % 2 != 0):
                                        for i in range(n1, n2 + 1, 2):
                                            ad = re.sub('(\d+)号?楼?[-～－~至]+(\d+)号?楼?', '', ad)
                                            ad1 = ad + str(i) + '号'
                                            with open('loudong.txt', 'a+')as f:
                                                f.write(area + ' ' + cmnt + ' ' + ad1 + '\n')
                                    else:
                                        for i in range(n1, n2 + 1):
                                            ad = re.sub('(\d+)号?楼?[-～－~至]+(\d+)号?楼?', '', ad)
                                            ad1 = ad + str(i) + '号'
                                            with open('loudong.txt', 'a+')as f:
                                                f.write(area + ' ' + cmnt + ' ' + ad1 + '\n')
                                else:
                                    with open('loudong.txt', 'a+')as f:
                                        f.write(area + ' ' + cmnt + ' ' + ad + '\n')


                            else:
                                with open('loudong.txt', 'a+')as f:
                                    f.write(area + ' ' + cmnt + ' ' + ad + '\n')
                        else:
                            ad = ad_pre + ad + '号'
                            address_reg1 = re.search('(\d+)号?楼?[-～－~至]+(\d+)号?楼?', ad)
                            address_reg2 = re.search('(\d+)号?楼?[-～－~至]+(\d+)号?楼?[-～－~至]+(\d+)号?楼?', ad)
                            if address_reg2:
                                with open('loudong.txt', 'a+')as f:
                                    f.write(area + ' ' + cmnt + ' ' + ad + '\n')

                            elif address_reg1:
                                n1 = int(address_reg1.group(1))
                                n2 = int(address_reg1.group(2))
                                if n1 < n2:
                                    if (n1 % 2 == 0 and n2 % 2 == 0) or (n1 % 2 != 0 and n2 % 2 != 0):
                                        for i in range(n1, n2 + 1, 2):
                                            ad = re.sub('(\d+)号?楼?[-～－~至]+(\d+)号?楼?', '', ad)
                                            ad1 = ad + str(i) + '号'
                                            with open('loudong.txt', 'a+')as f:
                                                f.write(area + ' ' + cmnt + ' ' + ad1 + '\n')
                                    else:
                                        for i in range(n1, n2 + 1):
                                            ad = re.sub('(\d+)号?楼?[-～－~至]+(\d+)号?楼?', '', ad)
                                            ad1 = ad + str(i) + '号'
                                            with open('loudong.txt', 'a+')as f:
                                                f.write(area + ' ' + cmnt + ' ' + ad1 + '\n')
                                else:
                                    with open('loudong.txt', 'a+')as f:
                                        f.write(area + ' ' + cmnt + ' ' + ad + '\n')


                            else:
                                with open('loudong.txt', 'a+')as f:
                                    f.write(area + ' ' + cmnt + ' ' + ad + '\n')

            elif str(address).find(';') == -1 or str(address).find('|') == -1 or str(address).find('.') == -1 or str(address).find('、') == -1:
                address_reg1 = re.search('(\d+)号?楼?[-～－~至]+(\d+)号?楼?', address)
                address_reg2 = re.search('(\d+)号?楼?[-～－~至]+(\d+)号?楼?[-～－~至]+(\d+)号?楼?', address)
                if address_reg2:
                    with open('loudong.txt', 'a+')as f:
                        f.write(area + ' ' + cmnt + ' ' + address + '\n')

                elif address_reg1:
                    n1 = int(address_reg1.group(1))
                    n2 = int(address_reg1.group(2))
                    if n1 < n2:
                        if (n1 % 2 == 0 and n2 % 2 == 0) or (n1 % 2 != 0 and n2 % 2 != 0):
                            for i in range(n1, n2 + 1, 2):
                                ad = re.sub('(\d+)号?楼?[-～－~至]+(\d+)号?楼?', '', address)
                                ad1 = ad + str(i) + '号'
                                with open('loudong.txt', 'a+')as f:
                                    f.write(area + ' ' + cmnt + ' ' + ad1 + '\n')
                        else:
                            for i in range(n1, n2 + 1):
                                ad = re.sub('(\d+)号?楼?[-～－~至]+(\d+)号?楼?', '', address)
                                ad1 = ad + str(i) + '号'
                                with open('loudong.txt', 'a+')as f:
                                    f.write(area + ' ' + cmnt + ' ' + ad1 + '\n')
                    else:
                        with open('loudong.txt', 'a+')as f:
                            f.write(area + ' ' + cmnt + ' ' + address + '\n')


                else:
                    with open('loudong.txt', 'a+')as f:
                        f.write(area + ' ' + cmnt + ' ' + address + '\n')
                # with open('loudong.txt', 'a+')as f:
                #     f.write(area + ' ' + cmnt + ' ' + address + '\n')


        except:
            address = '未知'
            with open('loudong.txt', 'a+')as f:
                f.write(area + ' ' + cmnt + ' ' + address + '\n')


if __name__ == '__main__':
    # read_cmnt()
    begin = time.time()
    print('程序开始，请等待。。。')
    read_list()
    print('程序结束, 数据写入到"loudong.txt中"')
    end = time.time()
    print('time is %d seconds ' % (end - begin))